import json
import openpyxl
from collections import defaultdict

non_metric_keys = ["id", "description", "ATTACK", "ENV", "ATTACK_VECTOR", "OUTCOME", "PRE_REQ"]
metric_keys = {"attackVector", "attackComplexity", "privilegesRequired", "userInteraction",
               "confidentialityImpact", "integrityImpact", "availabilityImpact", "impactScore", "baseSeverity",
               "accessVector", "accessComplexity", "authentication", "confidentialityImpact", "integrityImpact",
               "availabilityImpact", "baseSeverity"}

key_order = ["id", "description", "EXTRACTED_ATTACK", "EXTRACTED_ENV", "EXTRACTED_ATTACK_VECTOR", "EXTRACTED_OUTCOME", "EXTRACTED_PRE_REQ"]

def handle_non_metrics(data):
    extracted_keys = ["ATTACK", "ENV", "ATTACK_VECTOR", "OUTCOME", "PRE_REQ"]
    vuln_keys = data.keys()
    flat = {"EXTRACTED_" + k if k in extracted_keys else k: data[k] for k in non_metric_keys if k in vuln_keys}
    return flat

def handle_metrics(data, pref):
    metrics = {}
    for k in data.keys():
        if k not in metric_keys:
            continue
        metrics[pref + k] = data[k]
    for k in data["cvssData"].keys():
        if k not in metric_keys:
            continue
        metrics[pref + k] = data["cvssData"][k]
    return metrics

def flatten_vuln(vuln):
    flat = handle_non_metrics(vuln)
    metrics_keys = list(vuln["metrics"].keys())
    for metric in metrics_keys:
        data = vuln["metrics"][metric]
        flat = {**flat, **handle_metrics(data[0], metric + "_")}

    return flat


def convert_to_excel(vulns, output_path):
    data = [flatten_vuln(vuln) for vuln in vulns]
    all_keys = set()
    for d in data:
        all_keys.update(d.keys())
    new_key_order = key_order.copy()
    for k in all_keys:
        if k not in new_key_order:
            new_key_order.append(k)

    rows = defaultdict(lambda: ['NA'] * len(new_key_order))
    for d in data:
        row = rows[len(rows)]
        for i, key in enumerate(new_key_order):
            row[i] = d.get(key, 'NA')

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet.append(new_key_order)

    # Write the data rows
    for row_data in rows.values():
        worksheet.append(row_data)
    worksheet.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 24
    worksheet.column_dimensions[openpyxl.utils.get_column_letter(2)].width = 50
    worksheet.column_dimensions[openpyxl.utils.get_column_letter(3)].width = 30
    worksheet.column_dimensions[openpyxl.utils.get_column_letter(4)].width = 30
    worksheet.column_dimensions[openpyxl.utils.get_column_letter(5)].width = 30
    worksheet.column_dimensions[openpyxl.utils.get_column_letter(6)].width = 30
    worksheet.column_dimensions[openpyxl.utils.get_column_letter(7)].width = 30
    workbook.save(output_path)
    print(f"Excel file {output_path} created successfully.")


with open("./artifacts/final_bert_results.json", "r") as f: vulns_bert = json.load(f)
with open("./artifacts/final_crf_results.json", "r") as f: vulns_crf = json.load(f)
with open("./artifacts/final_hmm_results.json", "r") as f: vulns_hmm = json.load(f)


convert_to_excel(vulns_bert, "./artifacts/final_bert_results.xlsx")
convert_to_excel(vulns_crf, "./artifacts/final_crf_results.xlsx")
convert_to_excel(vulns_hmm, "./artifacts/final_hmm_results.xlsx")